import columns

def Scraping(product):
    from selenium import webdriver

    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument('headless')

    productToSearch = product.replace(' ', '+')

    driver = webdriver.Chrome()
    search = 'https://www.amazon.com.br/s?k=' + str(productToSearch)
    driver.get(search)

    results = driver.find_element_by_css_selector('.s-widget-spacing-small')
    firstProduct = results.find_element_by_css_selector('.sg-col-inner')

    wholePrice = firstProduct.find_element_by_css_selector('.a-price-whole').text
    fractionaryPrice = firstProduct.find_element_by_css_selector('.a-price-fraction').text
    allPrice = float(str(wholePrice).replace('.', '') + '.' + str(fractionaryPrice))

    return product, allPrice


def creatingExcel(product, allPrice):
    from pathlib import Path
    from openpyxl import Workbook
    import pandas as pd

    try:
        Path(r'C:/Users/Henrique/Desktop/Folder').mkdir()  # Make sure there are a folder
    except:
        pass

    folder = Path(r'C:/Users/Henrique/Desktop/Folder').iterdir()

    firstProductWord = product.split()[0]
    exist = False
    path = r'C:\Users\Henrique\Desktop\Folder\REPLACE HERE.xlsx'.replace('REPLACE HERE', ProductName)
    for file in folder:
        if str(file) == path:
            exist = True
            break

    if exist == False:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=ProductName)
        ws.cell(row=2, column=1, value=allPrice)
        wb.save(path)

    else:
        excelDF = pd.read_excel(path)
        newIndex = excelDF.index.max() + 1

        excelDF.loc[newIndex, ProductName] = allPrice

        arrayPrices = excelDF[ProductName].values
        columnPrices = list(arrayPrices)

        # _________DEFINING_THE_COLUMNS_________ #
        # Use this pattern

        columnVariation = columns.FunctionVariation(columnPrices)  # A funcion to get the values
        excelDF['Variation'] = columnVariation  # Add the column to the Data Frame, and choose a name
        excelDF.to_excel(path, index=False)

        # If you continue this pattern, you can put as many column as you want